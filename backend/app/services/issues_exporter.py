"""
Issues Exporter — Export contract analysis issues to Excel/CSV.

Phase 7: Output format support for structured issue export.
"""

import io
import csv
import logging
from typing import List, Dict, Any
from datetime import datetime, timezone

logger = logging.getLogger(__name__)


def export_issues_to_csv(issues: List[Dict[str, Any]], document_name: str = "contract") -> bytes:
    """Export analysis issues to CSV format.

    Args:
        issues: List of issue dicts with keys like clause_text, risk_level,
                recommendation, category, confidence, etc.
        document_name: Name of the source document.

    Returns:
        CSV file content as bytes.
    """
    output = io.StringIO()
    fieldnames = [
        "issue_number",
        "clause_text",
        "risk_level",
        "category",
        "recommendation",
        "suggested_revision",
        "confidence",
        "section",
        "page",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for idx, issue in enumerate(issues, 1):
        row = {
            "issue_number": idx,
            "clause_text": issue.get("clause_text", ""),
            "risk_level": issue.get("risk_level", ""),
            "category": issue.get("category", ""),
            "recommendation": issue.get("recommendation", ""),
            "suggested_revision": issue.get("suggested_revision", ""),
            "confidence": issue.get("confidence", ""),
            "section": issue.get("section", ""),
            "page": issue.get("page", ""),
        }
        writer.writerow(row)

    return output.getvalue().encode("utf-8")


def export_issues_to_xlsx(issues: List[Dict[str, Any]], document_name: str = "contract") -> bytes:
    """Export analysis issues to Excel (XLSX) format.

    Args:
        issues: List of issue dicts.
        document_name: Name of the source document.

    Returns:
        XLSX file content as bytes.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        logger.error("openpyxl not installed — Excel export unavailable")
        raise RuntimeError("Excel export requires openpyxl. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Contract Issues"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "#", "Clause Text", "Risk Level", "Category",
        "Recommendation", "Suggested Revision", "Confidence", "Section",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Risk level color mapping
    risk_fills = {
        "critical": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "high": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "medium": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "low": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }

    for idx, issue in enumerate(issues, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=issue.get("clause_text", "")).border = thin_border
        risk_cell = ws.cell(row=row, column=3, value=issue.get("risk_level", ""))
        risk_cell.border = thin_border
        risk_level = issue.get("risk_level", "").lower()
        if risk_level in risk_fills:
            risk_cell.fill = risk_fills[risk_level]
        ws.cell(row=row, column=4, value=issue.get("category", "")).border = thin_border
        ws.cell(row=row, column=5, value=issue.get("recommendation", "")).border = thin_border
        ws.cell(row=row, column=6, value=issue.get("suggested_revision", "")).border = thin_border
        conf = issue.get("confidence")
        ws.cell(row=row, column=7, value=f"{conf:.0%}" if isinstance(conf, (int, float)) else str(conf or "")).border = thin_border
        ws.cell(row=row, column=8, value=issue.get("section", "")).border = thin_border

    # Column widths
    col_widths = [5, 50, 12, 15, 40, 40, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Metadata row
    ws2 = wb.create_sheet("Metadata")
    ws2.cell(row=1, column=1, value="Document")
    ws2.cell(row=1, column=2, value=document_name)
    ws2.cell(row=2, column=1, value="Exported At")
    ws2.cell(row=2, column=2, value=datetime.now(timezone.utc).isoformat())
    ws2.cell(row=3, column=1, value="Total Issues")
    ws2.cell(row=3, column=2, value=len(issues))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
